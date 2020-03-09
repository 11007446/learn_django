#from django.shortcuts import render
import xlrd
import uuid
import time

from datetime import datetime

from django.http import HttpResponse
from django.utils.http import urlquote
from django.template import loader
from django.shortcuts import redirect
from django.db.models import Q
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage, InvalidPage
from django.http.response import JsonResponse

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font, colors, Alignment, Border, Side, PatternFill
#from openpyxl.utils import get_column_letter

from .models import Meeting
from . import tests
from . import mailsender

# Create your views here.


def exportdatalist(request):
    meetings = None
    # if "POST" == request.method:  # 条件查询
    concat = request.GET
    # concat['m_room_s'] 如果key对应的值为空，会抛出Keyerror错误
    # concat.get('key') 如果key对应的值为空，返回 None

    # 条件查询
    q1 = Q()
    q1.connector = 'AND'
    preparQ(q1, 'm_room', concat.get('m_room_s'))
    preparQ(q1, 'm_lotno', concat.get('m_lotno_s'), con_suf='__contains')
    preparQ(q1, 'm_name', concat.get('m_name_s'), con_suf='__contains')
    preparQ(q1, 'm_guide', concat.get('m_guide_s'), con_suf='__contains')
    preparQ(q1, 'm_mp', concat.get('m_mp_s'), con_suf='__contains')
    preparQ(q1, 'm_date', concat.get('m_date1'), con_suf='__gte')
    preparQ(q1, 'm_date', concat.get('m_date2'), con_suf='__lte')
    preparQ(q1, 'm_symbol', concat.get('m_symbol_s'), con_suf='__contains')

    meetings = Meeting.objects.filter(q1).order_by('m_date', 'm_stime')

    # 组织生成Excel内容
    book = Workbook()
    ws = book.active
    ws.title = '数据导出'
    # 导出表格标题
    ws['A1'] = '序号'
    ws['B1'] = '答辩地点'
    ws['C1'] = '答辩日期'
    ws['D1'] = '答辩时间'
    ws['E1'] = '项目名称'
    ws['F1'] = '指南名称'
    ws['G1'] = '单位'
    ws['H1'] = '项目负责人'
    ws['I1'] = '联系方式'

    rowindex = 2
    for meeting in meetings:
        ws.cell(rowindex, 1, (rowindex - 1))
        ws.cell(rowindex, 2, meeting.m_room)
        ws.cell(rowindex, 3, (meeting.m_date.strftime('%Y-%m-%d')))
        ws.cell(rowindex, 4, (meeting.m_stime.strftime("%H:%M") + " - " +
                              meeting.m_etime.strftime("%H:%M")))
        ws.cell(rowindex, 5, meeting.m_name)
        ws.cell(rowindex, 6, meeting.m_guide)
        ws.cell(rowindex, 7, meeting.m_org)
        ws.cell(rowindex, 8, meeting.m_mp)
        ws.cell(rowindex, 9, meeting.m_mobile)

        rowindex = rowindex + 1
        pass

    #调整表格列宽 中文效果不佳 弃用
    # column_widths = []
    # for row in ws:
    #     for i, cell in enumerate(row):
    #         lencellvalue = len(str(cell.value))
    #         if len(column_widths) > i:
    #             if lencellvalue > column_widths[i]:
    #                 column_widths[i] = lencellvalue
    #         else:
    #             column_widths += [lencellvalue]

    # for i, column_width in enumerate(column_widths):
    #     ws.column_dimensions[get_column_letter(i + 1)].width = column_width + 5

    # 文件输出response
    response = HttpResponse(
        save_virtual_workbook(book),
        content_type=
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    disposition = ('attachment;filename={}').format('')
    # 使用ajax下载文件， 文件名由前端JS决定，所以这里设置文件名已无效，故设置空字窜
    response['Content-Disposition'] = disposition
    return response


def sendweekplan(request):
    weekdayrange = mailsender.getweekdayrange()
    q1 = Q()
    q1.connector = 'AND'
    preparQ(q1, 'm_date', weekdayrange[0], con_suf='__gte')
    preparQ(q1, 'm_date', weekdayrange[1], con_suf='__lte')
    #meetings = Meeting.objects.filter(q1).order_by('m_date', 'm_stime')
    meetings = Meeting.objects.order_by('m_date', 'm_stime')
    # 测试代码
    # result = mailsender.sendweekplanmail(meetings)
    # return JsonResponse({'code': '0'})
    if (len(meetings) > 0):
        result = mailsender.sendweekplanmail(meetings)
        if (result):
            return JsonResponse({'code': '0'})
        else:
            return JsonResponse({'code': '1'})
    else:
        return JsonResponse({'code': '2'})


def cleardata(request):
    tests.cleardata()
    return HttpResponse('数据已清空！')


def formatcelldate(datestr):
    """[格式化excel单元格日期值]
    预留该函数应对日后各种日期格式的情况
    目前只针对 XXXX年XX月XX日 格式
    Arguments:
        datestr {[字符串]} -- [description]
    Returns:
        [datetime] -- [description]
    """
    if ('年' in datestr):
        formatresult = datetime.strptime(datestr, '%Y年%m月%d日')
    elif ('-' in datestr):
        formatresult = datetime.strptime(datestr, '%Y-%m-%d')
    elif ('/' in datestr):
        formatresult = datetime.strptime(datestr, '%Y/%m/%d')
    else:
        formatresult = ""
    return formatresult


def formatcelltext(textstr):
    """[格式化excel单元格文本值,去除空格、制表符、回车符、换行符]
    Arguments:
        textstr {[type]} -- [description]
    """
    formattext = ''
    if (textstr):
        formattext = str(textstr).replace(" ", "").replace("\t", "").replace(
            '\n', ' ').replace('\r', ' ')
    return formattext


def preparQ(q, con_name, con_value, con_suf='', func_format=None):
    """将查询条件填充Q查询对象
    Arguments:
        q {[type]} -- [Q查询对象]
        con_suf {[type]} -- [查询条件后缀 例: __contains 包含 __gte 大等于 __lte 小等于 ] (default: "")
        con_name {[type]} -- [查询参数名称]
        con_value {[type]} -- [查询参数值]
    Keyword Arguments:
        func_format {[type]} -- [格式化参数值回调函数] (default: {None})
    Returns:
        [type] -- [返回填充后的Q对象]
    """
    if (con_value):  # 参数值非空,该查询条件起效
        if (func_format):  # 提供参数格式化函数,对参数进行格式化
            q.children.append((con_name + con_suf, func_format(con_value)))
        else:
            q.children.append((con_name + con_suf, con_value))
    return q


def delcheckedmeetingdata(request):
    if (request.method == 'POST'):
        concat = request.POST
        pids = concat['checkedPid']
        if (pids):
            Meeting.objects.filter(pid__in=pids.split(',')).delete()
            datafilter()
    reponsestr = '{"code":"0"}'
    return HttpResponse(reponsestr)


def indexlayui(request):
    template = loader.get_template('meetingmanage/indexnew.htm')
    context = {}
    return HttpResponse(template.render(context, request))


def indexnewquery(request):
    meetings = None
    # if "POST" == request.method:  # 条件查询
    concat = request.GET
    # concat['m_room_s'] 如果key对应的值为空，会抛出Keyerror错误
    # concat.get('key') 如果key对应的值为空，返回 None

    # 条件查询
    q1 = Q()
    q1.connector = 'AND'
    preparQ(q1, 'm_room', concat.get('m_room_s'))
    preparQ(q1, 'm_lotno', concat.get('m_lotno_s'), con_suf='__contains')
    preparQ(q1, 'm_name', concat.get('m_name_s'), con_suf='__contains')
    preparQ(q1, 'm_guide', concat.get('m_guide_s'), con_suf='__contains')
    preparQ(q1, 'm_mp', concat.get('m_mp_s'), con_suf='__contains')
    preparQ(q1, 'm_date', concat.get('m_date1'), con_suf='__gte')
    preparQ(q1, 'm_date', concat.get('m_date2'), con_suf='__lte')
    preparQ(q1, 'm_symbol', concat.get('m_symbol_s'), con_suf='__contains')

    meetings = Meeting.objects.filter(q1).order_by('m_date', 'm_stime')
    # 分页处理
    page = concat.get('page')  # 当前页码
    limit_str = concat.get('limit')  # 当前记录/页
    limit = 10
    if (limit_str and limit_str.isdigit()):
        limit = int(limit_str)
    try:
        paginator = Paginator(meetings, limit)
        meetingpage = paginator.page(page)

    except PageNotAnInteger:
        # 如果请求的页数不是整数, 返回第一页。
        meetingpage = paginator.page(1)
    except InvalidPage:
        # 如果请求的页数不存在, 返回结果的最后一页。
        meetingpage = paginator.page(paginator.num_pages)
    except EmptyPage:
        # 如果请求的页数不在合法的页数范围内，返回结果的最后一页。
        meetingpage = paginator.page(paginator.num_pages)

    # layui 数据表格数据源格式化
    array_json = []
    jsonstr = ""
    allmeetingcount = len(meetings)
    for meeting in meetingpage:
        array_json.append(
            '{{"pid": "{}", "m_guide": "{}", "m_lotno": "{}", "m_room": "{}", "m_name": "{}", "m_date": "{}","m_time": "{}", "m_symbol": "{}", "m_mp": "{}", "m_mobile": "{}","m_org": "{}", "m_stime": "{}", "m_etime": "{}", "createtime": "{}"}}'
            .format(meeting.pid, meeting.m_guide, meeting.m_lotno,
                    meeting.m_room, meeting.m_name, meeting.m_date,
                    (meeting.m_stime.strftime("%H:%M") + " - " +
                     meeting.m_etime.strftime("%H:%M")), meeting.m_symbol,
                    meeting.m_mp, meeting.m_mobile, meeting.m_org,
                    meeting.m_stime, meeting.m_etime,
                    meeting.createtime.strftime('%Y-%m-%d %H:%M')))
    jsonstr = '{{"code":0,"msg":"","count":{},"data":[{}]}}'.format(
        allmeetingcount, ','.join(array_json))
    return HttpResponse(jsonstr)


def submitedit(request):
    if "POST" == request.method:
        concat = request.POST
        editPid = concat['editPid']
        meeting = Meeting.objects.get(pid=editPid)
        m_stime = concat['m_stime']
        m_etime = concat['m_etime']
        stime = datetime.strptime(m_stime.strip(), "%H:%M:%S")
        etime = datetime.strptime(m_etime.strip(), "%H:%M:%S")
        count_inteval = (etime - stime).seconds / 60
        meeting.m_room = concat['m_room']
        meeting.m_date = concat['m_date']
        meeting.m_guide = concat['m_guide']
        meeting.m_name = concat['m_name']
        meeting.m_mp = concat['m_mp']
        meeting.m_mobile = concat['m_mobile']
        meeting.m_stime = m_stime
        meeting.m_etime = m_etime
        meeting.m_inteval = count_inteval
        meeting.m_org = concat['m_org']
        meeting.save()
        datafilter()
    return HttpResponse('{"code":"0"}')


def importExcel(request):
    if "GET" == request.method:
        return redirect("/meetingmanage/")
    elif "POST" == request.method:
        wb = None
        try:
            concat = request.POST
            m_lotno = concat['m_lotno']
            excel_file = request.FILES['importExcel']
            wb = xlrd.open_workbook(filename=None,
                                    file_contents=excel_file.read())
            table = wb.sheets()[0]
            total_rows = table.nrows  # 拿到总共行数
            now = datetime.now()
            # 数据导入
            for rowindex in range(1, total_rows):  # 跳过首行标题
                row = table.row_values(rowindex)
                timestr = formatcelltext(row[2])
                stime = None
                etime = None
                interval = 0
                #FIXME 答辩时间解析, 间隔计算抽取到工具方法中
                if timestr.strip() != '':
                    timearray = timestr.split('-')
                    stime = datetime.strptime(timearray[0].strip(), "%H:%M")
                    etime = datetime.strptime(timearray[1].strip(), "%H:%M")
                    interval = (etime - stime).seconds / 60
                # 0答辩地点    1答辩日期    2答辩时间    3所属专项    4项目编号    5项目名称    6项目负责人    7联系方式    8申报单位    9推荐单位
                # print(row)
                pidstr = uuid.uuid4()
                meeting = Meeting(
                    pid=pidstr,
                    m_lotno=m_lotno,
                    m_room=formatcelltext(row[0]),
                    m_date=formatcelldate(row[1]),
                    m_guide=formatcelltext(row[3]),
                    m_number=formatcelltext(row[4]),
                    m_name=formatcelltext(row[5]),
                    m_mp=formatcelltext(row[6]),
                    m_mobile=formatcelltext(row[7]),
                    m_org=formatcelltext(row[8]),
                    m_orgre=formatcelltext(row[9]),
                    createtime=now,
                    m_stime=stime,
                    m_etime=etime,
                    m_inteval=interval,
                )
                # print(meeting)
                meeting.save()
            datafilter()
            wb.release_resources()

        except RuntimeError as e:
            print("导入出错: {0}".format(e))
            return HttpResponse('{"code":"1"}')
        finally:
            del wb
    return HttpResponse('{"code":"0"}')


def omitname(name, omitcount=5):
    if (len(name) > omitcount):
        return name[0:omitcount]
    else:
        return name


def datafilter():
    """过滤答辩数据,标识 冲突 连场 并场记录
    """
    #TODO 遍历集合
    #TODO 答辩日期相同、答辩室相同、起始时间相同 标识为冲突
    #TODO 答辩日期相同、答辩室相同、结束时间与其他答辩开始时间重合、标识为连场
    #TODO 答辩日期相同、答辩室不同、结束时间与其他答辩开始时间重合标识为并场
    q1 = Q()
    preparQ(q1,
            'm_date',
            time.strftime('%Y-%m-%d', time.localtime(time.time())),
            con_suf='__gte')
    meetingstofilter = Meeting.objects.filter(q1).order_by('m_date', 'm_stime')
    for meeting in meetingstofilter:
        meeting.m_symbol = ""
        for meeting_c in meetingstofilter:
            meeting_c.m_symbol = ""
            if (meeting.pid == meeting_c.pid):
                continue
            if (meeting.m_date == meeting_c.m_date):
                if (meeting.m_stime == meeting_c.m_stime):
                    if (meeting.m_room == meeting_c.m_room):
                        meeting.m_symbol = meeting.m_symbol + '与项目 [{}...] 冲突 '.format(
                            omitname(meeting_c.m_name))
                        meeting_c.m_symbol = meeting_c.m_symbol + '与项目 [{}...] 冲突 '.format(
                            omitname(meeting.m_name))
                    else:
                        meeting.m_symbol = meeting.m_symbol + '与项目 [{}...] 并场 '.format(
                            omitname(meeting_c.m_name))
                        meeting_c.m_symbol = meeting_c.m_symbol + '与项目 [{}...] 并场 '.format(
                            omitname(meeting.m_name))
                    meeting.save()
                    meeting_c.save()
                elif (meeting.m_etime == meeting_c.m_stime):
                    if (meeting.m_room == meeting_c.m_room):
                        meeting.m_symbol = meeting.m_symbol + '与项目 [{}...] 连场 '.format(
                            omitname(meeting_c.m_name))
                        meeting_c.m_symbol = meeting_c.m_symbol + '与项目 [{}...] 连场 '.format(
                            omitname(meeting.m_name))
                        meeting.save()
                        meeting_c.save()


def insertemptyrow(ws, rowcount, rowindex, border, rowheight=30):
    for index in range(rowcount):
        ws['A{}'.format(rowindex + index)] = index + 1
        ws['A{}'.format(rowindex + index)].alignment = Alignment(
            horizontal='center', vertical='center', wrapText=True)
        ws['A{}'.format(rowindex + index)].border = border
        ws['B{}'.format(rowindex + index)] = ''
        ws['B{}'.format(rowindex + index)].border = border
        ws['C{}'.format(rowindex + index)] = ''
        ws['C{}'.format(rowindex + index)].border = border
        ws['D{}'.format(rowindex + index)] = ''
        ws['D{}'.format(rowindex + index)].border = border
        ws.row_dimensions[(rowindex + index)].height = rowheight


def downloadsigninsheet(request):
    # 读取数据
    concat = request.GET
    m_lotno = concat['m_lotno']
    q1 = Q()
    q1.children.append(('m_lotno', m_lotno))
    meetings = Meeting.objects.filter(q1).order_by('m_room', 'm_date',
                                                   'm_stime')
    # 组织生成Excel内容

    # 使用ajax下载文件， 文件名由前端JS决定，所以这里设置已无效
    the_file_name = '批次{}_签到表_{}.xlsx'.format(
        m_lotno, time.strftime('%Y%m%d%H%M%S', time.localtime(time.time())))
    book = Workbook()
    sheetindex = 1

    a_centeralignment = Alignment(horizontal='center',
                                  vertical='center',
                                  wrapText=True)
    v_centeralignment = Alignment(vertical='center')
    border = Border(top=Side(border_style='thin', color=colors.BLACK),
                    bottom=Side(border_style='thin', color=colors.BLACK),
                    left=Side(border_style='thin', color=colors.BLACK),
                    right=Side(border_style='thin', color=colors.BLACK))

    titlefont = Font(name='仿宋_GB2312', size=25, color=colors.BLACK, bold=True)
    subtitlefont = Font(name='仿宋_GB2312', size=14, color=colors.BLACK)
    tabletitlefont = Font(name='仿宋_GB2312', size=12, color=colors.BLACK)
    fill = PatternFill("solid", fgColor='C9C9C9')

    for meeting in meetings:
        meetingdate = meeting.m_date.strftime('%Y-%m-%d')
        meetingtime = meeting.m_stime.strftime(
            "%H:%M") + " - " + meeting.m_etime.strftime("%H:%M")
        meetroom = meeting.m_room
        sheetname = '{} {}({})'.format(meetroom, meetingdate, sheetindex)
        if (sheetindex == 1):
            ws = book.active
            ws.title = sheetname
        else:
            ws = book.create_sheet(sheetname)
        ws.merge_cells('A1:D1')  # 签到表标题
        ws.merge_cells('A2:D2')  # 签到表标题
        ws.merge_cells('A3:D3')  # 签到表标题
        ws.merge_cells('A4:D4')  # 签到表标题
        ws['A1'] = '视频答辩签到表'
        ws['A2'] = '项目名称：{}'.format(meeting.m_name)
        ws['A3'] = '答辩日期：{} ({})'.format(meetingdate, meetingtime)
        ws['A4'] = '答辩地点：{}'.format(meeting.m_room)
        ws['A5'] = '序号'
        ws['B5'] = '项目负责人\n(主报告人)'
        ws['C5'] = '单位'
        ws['D5'] = '联系方式'
        insertemptyrow(ws, 3, 6, border)
        ws['A9'] = '序号'
        ws['B9'] = '其他参会人'
        ws['C9'] = '单位'
        ws['D9'] = '联系方式'
        insertemptyrow(ws, 15, 10, border)
        # 设置单元格格式

        titlerowheight = 45
        subtitlerowheight = 40
        tabletitlerowheight = 35
        ws['A1'].font = titlefont
        ws['A1'].alignment = a_centeralignment
        ws['A2'].font = subtitlefont
        ws['A2'].alignment = v_centeralignment
        ws['A3'].font = subtitlefont
        ws['A3'].alignment = v_centeralignment
        ws['A4'].font = subtitlefont
        ws['A4'].alignment = v_centeralignment

        ws['A5'].alignment = a_centeralignment
        ws['B5'].alignment = a_centeralignment
        ws['C5'].alignment = a_centeralignment
        ws['D5'].alignment = a_centeralignment
        ws['A5'].border = border
        ws['B5'].border = border
        ws['C5'].border = border
        ws['D5'].border = border
        ws['A5'].fill = fill
        ws['B5'].fill = fill
        ws['C5'].fill = fill
        ws['D5'].fill = fill

        ws['A9'].alignment = a_centeralignment
        ws['B9'].alignment = a_centeralignment
        ws['C9'].alignment = a_centeralignment
        ws['D9'].alignment = a_centeralignment
        ws['A9'].border = border
        ws['B9'].border = border
        ws['C9'].border = border
        ws['D9'].border = border
        ws['A9'].fill = fill
        ws['B9'].fill = fill
        ws['C9'].fill = fill
        ws['D9'].fill = fill

        ws.row_dimensions[1].height = titlerowheight
        ws.row_dimensions[2].height = subtitlerowheight
        ws.row_dimensions[3].height = subtitlerowheight
        ws.row_dimensions[4].height = subtitlerowheight
        ws.row_dimensions[5].height = tabletitlerowheight
        ws.row_dimensions[9].height = tabletitlerowheight
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 34
        ws.column_dimensions['D'].width = 22

        sheetindex = sheetindex + 1
        pass
    # 文件输出response
    response = HttpResponse(
        save_virtual_workbook(book),
        content_type=
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    the_file_name = urlquote(the_file_name)
    disposition = ('attachment;filename={}').format(the_file_name)
    response['Content-Disposition'] = disposition
    return response
